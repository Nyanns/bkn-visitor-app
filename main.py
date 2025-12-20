# FILE: backend/main.py (SECURE WITH JWT + LOGGING)
import shutil
import os
import uuid
import pytz
import zipfile
import io
from datetime import datetime, timedelta
from typing import List, Optional

from fastapi import FastAPI, Depends, HTTPException, File, UploadFile, Form, status, Request, BackgroundTasks, Response, Cookie
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.responses import FileResponse
from starlette.middleware.base import BaseHTTPMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import text, func, extract, desc, cast, Date
from pydantic import BaseModel
import bcrypt
from jose import JWTError, jwt
from dotenv import load_dotenv
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
import puremagic
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Setup Logging
os.makedirs("logs", exist_ok=True)
logger.add(
    "logs/app.log",
    rotation="500 MB",
    retention="7 days",
    format="{time:YYYY-MM-DD HH:mm:ss} | {level} | {message}",
    level="INFO"
)

# Load Environment Variables
load_dotenv()

# Import file lokal
from database import engine, get_db
import models

# Inisialisasi Database
models.Base.metadata.create_all(bind=engine)

# --- 1. KONFIGURASI KEAMANAN (JWT) ---
SECRET_KEY = os.getenv("SECRET_KEY")
if not SECRET_KEY:
    # Error kalau lupa bikin .env
    raise ValueError("FATAL: SECRET_KEY belum disetting di file .env!")

ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")

# --- 2. KONFIGURASI UTAMA ---
JAKARTA_TZ = pytz.timezone('Asia/Jakarta')
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB for task letters
MAX_TASK_LETTERS_PER_VISIT = 5  # Maximum surat tugas per visit
UPLOAD_DIR = "uploads"
TASK_LETTER_DIR = "uploads/task_letters"  # Dedicated folder for surat tugas
BACKUP_DIR = "backups"
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(TASK_LETTER_DIR, exist_ok=True)
os.makedirs(BACKUP_DIR, exist_ok=True)

# Setup Limiter
limiter = Limiter(key_func=get_remote_address)

# Initialize FastAPI with API Documentation
app = FastAPI(
    title="BKN Visitor System",
    version="1.0.0",
    description="Sistem Manajemen Tamu BKN - Direktorat INTIKAMI",
    docs_url="/api/docs",      # Swagger UI
    redoc_url="/api/redoc",    # ReDoc (alternative docs)
    openapi_url="/api/openapi.json"
)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# SECURITY: Public access removed - use protected endpoint instead
# app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")

# --- 3. MIDDLEWARE & SECURITY ---

# A. PROXY HEADERS (For Nginx/Docker) - Add this if behind proxy
# from uvicorn.middleware.proxy_headers import ProxyHeadersMiddleware
# app.add_middleware(ProxyHeadersMiddleware, trusted_hosts="*")

# B. SECURE HEADERS (Inner Middleware)
# Using standard middleware function instead of BaseHTTPMiddleware to avoid FileResponse issues
@app.middleware("http")
async def secure_headers_middleware(request: Request, call_next):
    try:
        response = await call_next(request)
        # Basic Security Headers
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["X-XSS-Protection"] = "1; mode=block"
        response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
        return response
    except Exception as e:
        # If an error occurs during response generation (e.g. streaming), re-raise
        # so FastAPI error handlers can catch it, or handle it here.
        # For now, let it propagate so CORS (Outer) can handle the error response if any.
        logger.error(f"Middleware error: {e}")
        raise e

# C. CORS (Outer Middleware - Added LAST to wrap everything)
# Load allowed origins from environment variable (comma-separated)
env_origins = os.getenv("ALLOWED_ORIGINS", "")
default_origins = [
    "http://localhost:5173",
    "http://localhost:5174",
    "http://127.0.0.1:5173", 
    "http://127.0.0.1:5174"
]

if env_origins:
    ALLOWED_ORIGINS = env_origins.split(",")
    # Combine with defaults to ensure dev environment always works
    for origin in default_origins:
        if origin not in ALLOWED_ORIGINS:
            ALLOWED_ORIGINS.append(origin)
else:
    ALLOWED_ORIGINS = default_origins

app.add_middleware(
    CORSMiddleware, 
    allow_origins=ALLOWED_ORIGINS, 
    allow_credentials=True, 
    allow_methods=["*"], 
    allow_headers=["*"]
)

# --- 4. SCHEMAS ---
class StandardResponse(BaseModel):
    status: str
    message: str
    data: Optional[dict] = None

class CheckInOutResponse(BaseModel):
    status: str
    message: str
    time: str

class Token(BaseModel):
    access_token: str
    token_type: str

# --- 5. FUNGSI KEAMANAN (HELPER) ---

def verify_password(plain_password, hashed_password):
    if isinstance(hashed_password, str):
        hashed_password = hashed_password.encode('utf-8')
    return bcrypt.checkpw(plain_password.encode('utf-8'), hashed_password)

def get_password_hash(password):
    salt = bcrypt.gensalt()
    hashed = bcrypt.hashpw(password.encode('utf-8'), salt)
    return hashed.decode('utf-8')

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=15))
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

# --- DEPENDENCY: PENJAGA PINTU (Cek Token) ---
async def get_current_admin(request: Request, db: Session = Depends(get_db)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Login kadaluarsa atau tidak valid",
        headers={"WWW-Authenticate": "Bearer"},
    )
    
    # 1. Try Cookie
    token = request.cookies.get("access_token")
    
    # 2. Try Header (Fallback/Swagger)
    if not token:
        auth_header = request.headers.get("Authorization")
        if auth_header and auth_header.startswith("Bearer "):
            token = auth_header.split(" ")[1]

    if not token:
        raise credentials_exception

    try:
        # Check if "Bearer " prefix is in cookie value (legacy check)
        if token.startswith("Bearer "):
            token = token.split(" ")[1]
            
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
        
    admin = db.query(models.Admin).filter(models.Admin.username == username).first()
    if admin is None:
        raise credentials_exception
    return admin

# --- 6. FUNGSI VALIDASI LAINNYA ---
def validate_nik(nik: str):
    if not nik.isdigit():
        raise HTTPException(status_code=400, detail="Format NIK salah! Harus angka.")
    return nik

def validate_and_save_file(upload_file: UploadFile) -> str:
    if not upload_file: 
        raise HTTPException(status_code=400, detail="File wajib!")
    
    file_ext = os.path.splitext(upload_file.filename)[1].lower()
    ALLOWED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".pdf"}
    if file_ext not in ALLOWED_EXTENSIONS:
        logger.error(f"File upload rejected: invalid extension {file_ext}")
        raise HTTPException(status_code=400, detail=f"Ekstensi file tidak diizinkan: {file_ext}")

    try:
        header = upload_file.file.read(2048)
        upload_file.file.seek(0)
        
        # Use puremagic to detect mime type from bytes
        # puremagic.from_string returns list, use magic_string for simple detection or try/except
        try:
             # puremagic.magic_string returns list of PureMagicResult
             # We want the highest confidence or first match
             detected_mimes = puremagic.magic_string(header)
             if not detected_mimes:
                 raise ValueError("Unknown file type")
             mime_type = detected_mimes[0].mime_type
        except Exception:
             mime_type = "application/octet-stream"
        
        ALLOWED_MIMES = {"image/jpeg", "image/png", "application/pdf"}
        if mime_type not in ALLOWED_MIMES:
            logger.error(f"File upload rejected: invalid MIME type {mime_type}")
            raise HTTPException(status_code=400, detail=f"Tipe file tidak valid (terdeteksi: {mime_type})")
             
    except Exception as e:
        logger.error(f"Error validating file: {e}")
        raise HTTPException(status_code=400, detail="Gagal memvalidasi file.")

    unique_name = f"{uuid.uuid4()}{file_ext}"
    file_path = os.path.join(UPLOAD_DIR, unique_name)
    try:
        with open(file_path, "wb") as buffer: 
            shutil.copyfileobj(upload_file.file, buffer)
        logger.info(f"File uploaded successfully: {unique_name}")
    except Exception as e:
        logger.error(f"Failed to save file: {e}")
        raise HTTPException(500, "Gagal simpan file")
    return file_path


# --- 7. API ENDPOINTS ---

@app.get("/", tags=["System"])
def read_root():
    """Root endpoint - API Status"""
    return {"status": "online", "system": "BKN Visitor App v1.0.0", "docs": "/api/docs"}

# === MONITORING ENDPOINTS ===
@app.get("/health", tags=["Monitoring"])
def health_check(db: Session = Depends(get_db)):
    """
    Health check endpoint for monitoring tools (e.g., Docker, Kubernetes)
    Returns HTTP 200 if system is healthy, 503 otherwise
    """
    try:
        # Check database connectivity
        db.execute(text("SELECT 1"))
        
        return {
            "status": "healthy",
            "timestamp": datetime.now(JAKARTA_TZ).isoformat(),
            "checks": {
                "database": "ok",
                "api": "ok"
            }
        }
    except Exception as e:
        logger.error(f"Health check failed: {e}")
        raise HTTPException(
            status_code=503, 
            detail={"status": "unhealthy", "error": "Database connection failed"}
        )

@app.get("/metrics", tags=["Monitoring"])
def get_metrics(db: Session = Depends(get_db)):
    """
    Basic metrics endpoint for monitoring
    Returns counts of visitors, visits, and active sessions
    """
    try:
        total_visitors = db.query(models.Visitor).count()
        total_visits = db.query(models.VisitLog).count()
        
        # Active visits today
        today = datetime.now(JAKARTA_TZ).date()
        active_visits = db.query(models.VisitLog).filter(
            models.VisitLog.visit_date == today,
            models.VisitLog.check_out_time == None
        ).count()
        
        return {
            "timestamp": datetime.now(JAKARTA_TZ).isoformat(),
            "metrics": {
                "total_visitors": total_visitors,
                "total_visits": total_visits,
                "active_visits_today": active_visits
            }
        }
    except Exception as e:
        logger.error(f"Metrics collection failed: {e}")
        raise HTTPException(status_code=500, detail="Failed to collect metrics")

# === LOGIN ADMIN ===
@app.post("/token")
@limiter.limit("5/minute")
def login_for_access_token(response: Response, request: Request, form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    client_ip = request.client.host
    admin = db.query(models.Admin).filter(models.Admin.username == form_data.username).first()
    
    if not admin or not verify_password(form_data.password, admin.password_hash):
        logger.warning(f"Failed login attempt for user '{form_data.username}' from {client_ip}")
        raise HTTPException(status_code=400, detail="Username atau Password Salah")
    
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(data={"sub": admin.username}, expires_delta=access_token_expires)
    
    # Set HttpOnly Cookie
    response.set_cookie(
        key="access_token",
        value=f"{access_token}",
        httponly=True,
        max_age=ACCESS_TOKEN_EXPIRE_MINUTES * 60,
        expires=ACCESS_TOKEN_EXPIRE_MINUTES * 60,
        samesite="lax",
        secure=False  # Set True for HTTPS in production
        # domain="localhost" # Optional, usually auto-set
    )
    
    logger.info(f"Successful login: {admin.username} from {client_ip}")
    return {"status": "success", "message": "Login berhasil"}

@app.post("/logout", tags=["Auth"])
def logout(response: Response):
    response.delete_cookie(key="access_token")
    return {"status": "success", "message": "Logout berhasil"}

# === SETUP ADMIN ===
@app.post("/setup-admin")
def create_initial_admin(username: str = Form(...), password: str = Form(...), db: Session = Depends(get_db)):
    if os.getenv("ALLOW_SETUP_ADMIN") != "true":
        logger.warning("Attempt to access /setup-admin with feature disabled")
        raise HTTPException(status_code=403, detail="Fitur setup admin dinonaktifkan.")

    if db.query(models.Admin).first():
        raise HTTPException(status_code=400, detail="Admin sudah ada. Gunakan login.")

    try:
        hashed_password = get_password_hash(password)
        new_admin = models.Admin(username=username, password_hash=hashed_password)
        db.add(new_admin)
        db.commit()
        logger.info(f"New admin created: {username}")
        return {"status": "success", "message": f"Admin {username} berhasil dibuat"}
    except Exception as e:
        logger.error(f"Error creating admin: {e}")
        raise HTTPException(status_code=500, detail="Terjadi kesalahan server. Silakan coba lagi.")

# === HELPER: CLEANUP EXCEL FILES ===
def cleanup_excel_file(filepath: str):
    """
    Background task to delete Excel file after download.
    Prevents disk space accumulation from exported files.
    """
    try:
        if os.path.exists(filepath):
            os.remove(filepath)
            logger.info(f"Cleaned up Excel file: {filepath}")
    except Exception as e:
        logger.error(f"Error cleaning up Excel file {filepath}: {e}")

# === EXPORT TO EXCEL ===

# === EXPORT TO EXCEL ===
@app.get("/admin/export-excel", tags=["Admin"])
def export_to_excel(request: Request, background_tasks: BackgroundTasks, current_admin: models.Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    try:
        # Determine Base URL for Hyperlinks
        # Use env var if set (for production proxy), otherwise use request URL
        base_url = os.getenv("APP_BASE_URL")
        if not base_url:
            base_url = str(request.base_url).rstrip("/")
        
        # Query with eager loading for relationships
        results = db.query(models.VisitLog, models.Visitor)\
            .join(models.Visitor, models.VisitLog.visitor_nik == models.Visitor.nik)\
            .outerjoin(models.Room, models.VisitLog.room_id == models.Room.id)\
            .outerjoin(models.Companion, models.VisitLog.companion_id == models.Companion.id)\
            .order_by(models.VisitLog.check_in_time.desc()).all()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Tamu"
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Headers
        headers = [
            "No", "NIK", "Nama Lengkap", "Instansi", "No. HP", 
            "Tujuan", "Ruangan", "Pendamping",
            "Check-in", "Check-out", "Status",
            "Foto", "KTP", "Surat Tugas"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows with timezone conversion
        jakarta_tz = pytz.timezone('Asia/Jakarta')
        
        # Style for hyperlinks
        link_font = Font(color="0563C1", underline="single")
        
        for idx, (log, visitor) in enumerate(results, 2):
            ws.cell(row=idx, column=1, value=idx-1)
            ws.cell(row=idx, column=2, value=visitor.nik)
            ws.cell(row=idx, column=3, value=visitor.full_name)
            ws.cell(row=idx, column=4, value=visitor.institution)
            ws.cell(row=idx, column=5, value=visitor.phone or "-")
            ws.cell(row=idx, column=6, value=log.visit_purpose or "-")
            ws.cell(row=idx, column=7, value=log.room.name if log.room else "-")
            ws.cell(row=idx, column=8, value=log.companion.name if log.companion else "-")
            
            # Convert UTC to Jakarta timezone for display
            if log.check_in_time:
                check_in_utc = log.check_in_time.replace(tzinfo=pytz.UTC)
                check_in_jkt = check_in_utc.astimezone(jakarta_tz)
                ws.cell(row=idx, column=9, value=check_in_jkt.strftime("%Y-%m-%d %H:%M"))
            else:
                ws.cell(row=idx, column=9, value="-")
                
            if log.check_out_time:
                check_out_utc = log.check_out_time.replace(tzinfo=pytz.UTC)
                check_out_jkt = check_out_utc.astimezone(jakarta_tz)
                ws.cell(row=idx, column=10, value=check_out_jkt.strftime("%Y-%m-%d %H:%M"))
            else:
                ws.cell(row=idx, column=10, value="-")
                
            ws.cell(row=idx, column=11, value="Sedang Berkunjung" if not log.check_out_time else "Selesai")

            # === Document Statuses with HYPERLINKS ===
            
            # 1. FOTO
            cell_foto = ws.cell(row=idx, column=12)
            if visitor.photo_path and os.path.exists(visitor.photo_path):
                filename = os.path.basename(visitor.photo_path)
                file_url = f"{base_url}/uploads/{filename}"
                cell_foto.value = "Lihat Foto"
                cell_foto.hyperlink = file_url
                cell_foto.font = link_font
            else:
                cell_foto.value = "Tidak Ada"

            # 2. KTP
            cell_ktp = ws.cell(row=idx, column=13)
            if visitor.ktp_path and os.path.exists(visitor.ktp_path):
                filename = os.path.basename(visitor.ktp_path)
                file_url = f"{base_url}/uploads/{filename}"
                cell_ktp.value = "Lihat KTP"
                cell_ktp.hyperlink = file_url
                cell_ktp.font = link_font
            else:
                cell_ktp.value = "Tidak Ada"
            
            # 3. SURAT TUGAS
            cell_letter = ws.cell(row=idx, column=14)
            
            # Check for modern task letters first
            # (In a real scenario, could link to a zip or list multiple, for now linking first or legacy)
            letter_count = db.query(models.TaskLetter).filter(models.TaskLetter.visit_id == log.id).count()
            
            if letter_count > 0:
                if letter_count > 1:
                    # Link to ZIP download
                    file_url = f"{base_url}/visits/{log.id}/task-letters/archive"
                    cell_letter.value = f"Download {letter_count} Files (ZIP)"
                else:
                    # Link to the single file
                    first_letter = db.query(models.TaskLetter).filter(models.TaskLetter.visit_id == log.id).first()
                    clean_filename = os.path.basename(first_letter.file_path)
                    file_url = f"{base_url}/uploads/{clean_filename}"
                    cell_letter.value = f"1 File (Lihat)"
                
                cell_letter.hyperlink = file_url
                cell_letter.font = link_font
                
            elif visitor.task_letter_path and os.path.exists(visitor.task_letter_path):
                # Legacy
                filename = os.path.basename(visitor.task_letter_path)
                file_url = f"{base_url}/uploads/{filename}"
                cell_letter.value = "Lihat (Legacy)"
                cell_letter.hyperlink = file_url
                cell_letter.font = link_font
            else:
                cell_letter.value = "-"
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 18
        ws.column_dimensions['J'].width = 18
        ws.column_dimensions['K'].width = 18
        ws.column_dimensions['L'].width = 10
        ws.column_dimensions['M'].width = 10
        ws.column_dimensions['N'].width = 15
        
        # Save file
        # Save file with unique timestamp to prevent locking/permission issues
        filename = f"Laporan_Tamu_Lengkap_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        filepath = os.path.join(BACKUP_DIR, filename)
        wb.save(filepath)
        
        logger.info(f"Enhanced Excel exported by {current_admin.username}: {filename}")
        
        # Add background task to cleanup file after download
        background_tasks.add_task(cleanup_excel_file, filepath)
        
        return FileResponse(
            filepath, 
            filename=filename, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    except Exception as e:
        logger.error(f"Error exporting Excel: {e}")

# === EXPORT MASTER DATA ===
@app.get("/admin/export-master-data", tags=["Admin"])
def export_master_data(background_tasks: BackgroundTasks, current_admin: models.Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    try:
        # Create workbook
        wb = Workbook()
        
        # --- SHEET 1: ROOMS ---
        ws_rooms = wb.active
        ws_rooms.title = "Data Ruangan"
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Rooms Headers
        room_headers = ["No", "Nama Ruangan", "Deskripsi", "Status"]
        for col_num, header in enumerate(room_headers, 1):
            cell = ws_rooms.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            
        # Rooms Data
        rooms = db.query(models.Room).all()
        for idx, room in enumerate(rooms, 2):
            ws_rooms.cell(row=idx, column=1, value=idx-1)
            ws_rooms.cell(row=idx, column=2, value=room.name)
            ws_rooms.cell(row=idx, column=3, value=room.description or "-")
            ws_rooms.cell(row=idx, column=4, value="Aktif" if room.is_active else "Nonaktif")
            
        # Adjust column widths
        ws_rooms.column_dimensions['A'].width = 5
        ws_rooms.column_dimensions['B'].width = 30
        ws_rooms.column_dimensions['C'].width = 40
        ws_rooms.column_dimensions['D'].width = 15

        # --- SHEET 2: COMPANIONS ---
        ws_companions = wb.create_sheet(title="Data Pendamping")
        
        # Companions Headers
        comp_headers = ["No", "Nama Pendamping", "Jabatan", "Status"]
        for col_num, header in enumerate(comp_headers, 1):
            cell = ws_companions.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            
        # Companions Data
        companions = db.query(models.Companion).all()
        for idx, comp in enumerate(companions, 2):
            ws_companions.cell(row=idx, column=1, value=idx-1)
            ws_companions.cell(row=idx, column=2, value=comp.name)
            ws_companions.cell(row=idx, column=3, value=comp.position or "-")
            ws_companions.cell(row=idx, column=4, value="Aktif" if comp.is_active else "Nonaktif")
            
        # Adjust column widths
        ws_companions.column_dimensions['A'].width = 5
        ws_companions.column_dimensions['B'].width = 30
        ws_companions.column_dimensions['C'].width = 30
        ws_companions.column_dimensions['D'].width = 15
        
        # Save file
        filename = f"Master_Data_BKN_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        filepath = os.path.join(BACKUP_DIR, filename)
        wb.save(filepath)
        
        logger.info(f"Master Data exported by {current_admin.username}: {filename}")
        
        # Add background task to cleanup file after download
        background_tasks.add_task(cleanup_excel_file, filepath)
        
        return FileResponse(
            filepath, 
            filename=filename, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    except Exception as e:
        logger.error(f"Error exporting Master Data: {e}")
        raise HTTPException(status_code=500, detail="Gagal export Master Data")


# === PROTECTED ENDPOINTS ===
@app.get("/admin/logs", tags=["Admin"])
def get_admin_logs(current_admin: models.Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    results = db.query(models.VisitLog, models.Visitor)\
        .join(models.Visitor, models.VisitLog.visitor_nik == models.Visitor.nik)\
        .order_by(models.VisitLog.check_in_time.desc()).all()
    
    logs_data = []
    jakarta_tz = pytz.timezone('Asia/Jakarta')
    
    for log, visitor in results:
        status_visit = "Selesai" if log.check_out_time else "Sedang Berkunjung"
        photo_url = f"/uploads/{os.path.basename(visitor.photo_path)}" if visitor.photo_path else None
        
        # Convert UTC (stored as naive datetime) to Jakarta timezone
        if log.check_in_time:
            check_in_utc = log.check_in_time.replace(tzinfo=pytz.UTC)
            check_in_jkt = check_in_utc.astimezone(jakarta_tz)
        else:
            check_in_jkt = None
            
        if log.check_out_time:
            check_out_utc = log.check_out_time.replace(tzinfo=pytz.UTC)
            check_out_jkt = check_out_utc.astimezone(jakarta_tz)
        else:
            check_out_jkt = None
        
        logs_data.append({
            "id": log.id, 
            "nik": visitor.nik, 
            "full_name": visitor.full_name,
            "institution": visitor.institution, 
            "check_in_time": check_in_jkt,  # Send Jakarta time
            "check_out_time": check_out_jkt,  # Send Jakarta time
            "status": status_visit, 
            "photo_url": photo_url
        })
    return logs_data

@app.post("/visitors/", status_code=status.HTTP_201_CREATED, tags=["Admin"])
def create_visitor(
    nik: str = Form(..., min_length=3, max_length=20),
    full_name: str = Form(...), institution: str = Form(...), phone: str = Form(None),
    auto_checkin: bool = Form(False),  # New parameter
    ktp: UploadFile = File(...), photo: Optional[UploadFile] = File(None), task_letter: Optional[UploadFile] = File(None),
    current_admin: models.Admin = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    validate_nik(nik)
    if db.query(models.Visitor).filter(models.Visitor.nik == nik).first():
        raise HTTPException(400, f"NIK {nik} sudah terdaftar!")

    photo_path = validate_and_save_file(photo) if photo else None
    ktp_path = validate_and_save_file(ktp)
    task_letter_path = validate_and_save_file(task_letter) if task_letter else None

    # 1. Register Visitor
    new_visitor = models.Visitor(
        nik=nik, full_name=full_name.strip(), institution=institution.strip(),
        phone=phone, photo_path=photo_path, ktp_path=ktp_path, task_letter_path=task_letter_path
    )
    db.add(new_visitor)
    db.commit()

    # 2. Auto Check-In Logic
    message = "Registrasi Tamu Berhasil"
    if auto_checkin:
        now_jkt = datetime.now(JAKARTA_TZ)
        now_utc = now_jkt.astimezone(pytz.UTC).replace(tzinfo=None)
        
        new_log = models.VisitLog(
            visitor_nik=nik, 
            check_in_time=now_utc, 
            visit_date=now_jkt.date()
        )
        db.add(new_log)
        db.commit()
        message = "Registrasi & Auto Check-In Berhasil"
        logger.info(f"Auto Check-in: {full_name} (NIK: {nik}) at {now_jkt}")

    logger.info(f"New visitor registered: {full_name} (NIK: {nik}) by {current_admin.username}")
    return StandardResponse(status="success", message=message)

# === PUBLIC ENDPOINTS ===
@app.get("/visitors/{nik}", tags=["Visitor"])
def get_visitor(nik: str, db: Session = Depends(get_db)):
    visitor = db.query(models.Visitor).filter(models.Visitor.nik == nik).first()
    if not visitor: raise HTTPException(404, "Data tidak ditemukan")
    now_jkt = datetime.now(JAKARTA_TZ)
    active_visit = db.query(models.VisitLog).filter(
        models.VisitLog.visitor_nik == nik, models.VisitLog.visit_date == now_jkt.date(),
        models.VisitLog.check_out_time == None
    ).first()
    return {
        "nik": visitor.nik, 
        "full_name": visitor.full_name, 
        "institution": visitor.institution, 
        "phone": visitor.phone,
        "photo_path": visitor.photo_path, 
        "ktp_path": visitor.ktp_path,
        "task_letter_path": visitor.task_letter_path,
        "is_checked_in": active_visit is not None
    }

@app.get("/visitors/{nik}/photo", tags=["Visitor"])
def get_visitor_photo(nik: str, db: Session = Depends(get_db)):
    """
    Public endpoint for visitors to access their own photo.
    No admin authentication required - validates by NIK.
    """
    visitor = db.query(models.Visitor).filter(models.Visitor.nik == nik).first()
    if not visitor:
        raise HTTPException(status_code=404, detail="Visitor not found")
    
    if not visitor.photo_path:
        raise HTTPException(status_code=404, detail="Photo not found")
    
    # Check if file exists
    if not os.path.exists(visitor.photo_path):
        logger.warning(f"Photo file missing for visitor {nik}: {visitor.photo_path}")
        raise HTTPException(status_code=404, detail="Photo file not found")
    
    # Return photo with proper headers
    return FileResponse(
        visitor.photo_path,
        media_type="image/jpeg",
        headers={
            "Cache-Control": "public, max-age=86400",  # Cache for 24 hours
            "X-Content-Type-Options": "nosniff"
        }
    )


@app.post("/check-in/", tags=["Attendance"])
def check_in(
    nik: str = Form(...),
    visit_purpose: str = Form(None),  # NEW: Tujuan berkunjung (optional at check-in, can be added later)
    room_id: int = Form(None),  # NEW: Room ID
    companion_id: int = Form(None),  # NEW: Companion ID
    db: Session = Depends(get_db)
):
    validate_nik(nik)
    visitor = db.query(models.Visitor).filter(models.Visitor.nik == nik).first()
    if not visitor: raise HTTPException(404, "NIK tidak ditemukan.")
    
    # Validate room_id if provided
    if room_id:
        room = db.query(models.Room).filter(models.Room.id == room_id, models.Room.is_active == True).first()
        if not room:
            raise HTTPException(400, "Ruangan tidak valid atau tidak aktif.")
    
    # Validate companion_id if provided
    if companion_id:
        companion = db.query(models.Companion).filter(models.Companion.id == companion_id, models.Companion.is_active == True).first()
        if not companion:
            raise HTTPException(400, "Pendamping tidak valid atau tidak aktif.")
    
    # Get current time in Jakarta timezone
    now_jkt = datetime.now(JAKARTA_TZ)
    
    # Convert to UTC for storage (consistent timezone in database)
    now_utc = now_jkt.astimezone(pytz.UTC).replace(tzinfo=None)
    
    active_visit = db.query(models.VisitLog).filter(
        models.VisitLog.visitor_nik == nik, 
        models.VisitLog.visit_date == now_jkt.date(),
        models.VisitLog.check_out_time == None
    ).first()
    if active_visit: 
        return CheckInOutResponse(status="info", message="Sudah masuk.", time=now_jkt.strftime("%H:%M:%S"))
    
    # Store UTC time in database with visit details
    new_log = models.VisitLog(
        visitor_nik=nik, 
        check_in_time=now_utc, 
        visit_date=now_jkt.date(),
        visit_purpose=visit_purpose.strip() if visit_purpose else None,
        room_id=room_id,
        companion_id=companion_id
    )
    db.add(new_log)
    db.commit()
    db.refresh(new_log)
    
    logger.info(f"Check-in: {visitor.full_name} (NIK: {nik}) at {now_jkt.strftime('%Y-%m-%d %H:%M:%S %Z')}")
    return {
        "status": "success", 
        "message": f"Selamat Datang {visitor.full_name}!", 
        "time": now_jkt.strftime("%H:%M:%S"),
        "visit_id": new_log.id  # Return visit ID for file uploads
    }

@app.post("/check-out/", tags=["Attendance"])
def check_out(nik: str = Form(...), db: Session = Depends(get_db)):
    # Get current time in Jakarta timezone
    now_jkt = datetime.now(JAKARTA_TZ)
    
    # Convert to UTC for storage
    now_utc = now_jkt.astimezone(pytz.UTC).replace(tzinfo=None)
    
    active_visit = db.query(models.VisitLog).filter(
        models.VisitLog.visitor_nik == nik, 
        models.VisitLog.visit_date == now_jkt.date(),
        models.VisitLog.check_out_time == None
    ).first()
    if not active_visit: raise HTTPException(400, "Belum Check-In.")
    
    visitor = db.query(models.Visitor).filter(models.Visitor.nik == nik).first()
    
    # Store UTC time in database
    active_visit.check_out_time = now_utc
    db.commit()
    logger.info(f"Check-out: {visitor.full_name if visitor else 'Unknown'} (NIK: {nik}) at {now_jkt.strftime('%Y-%m-%d %H:%M:%S %Z')}")
    return CheckInOutResponse(status="success", message="Hati-hati di jalan!", time=now_jkt.strftime("%H:%M:%S"))

# === GET VISITOR HISTORY ===
@app.get("/visitors/{nik}/history", tags=["Visitor"])
def get_visitor_history(nik: str, db: Session = Depends(get_db)):
    """Get complete visit history for a visitor"""
    # Check if visitor exists
    visitor = db.query(models.Visitor).filter(models.Visitor.nik == nik).first()
    if not visitor:
        raise HTTPException(status_code=404, detail="Visitor not found")
    
    # Get all visit logs for this visitor, ordered by date descending
    logs = db.query(models.VisitLog)\
        .filter(models.VisitLog.visitor_nik == nik)\
        .order_by(models.VisitLog.check_in_time.desc())\
        .all()
    
    # Format history
    history = []
    jakarta_tz = pytz.timezone('Asia/Jakarta')
    
    for log in logs:
        # Database stores naive datetime (UTC)
        # Convert to Jakarta timezone for display
        if log.check_in_time:
            # Treat naive datetime as UTC, then convert to Jakarta
            check_in_utc = log.check_in_time.replace(tzinfo=pytz.UTC)
            check_in_jkt = check_in_utc.astimezone(jakarta_tz)
        else:
            check_in_jkt = None
            
        if log.check_out_time:
            check_out_utc = log.check_out_time.replace(tzinfo=pytz.UTC)
            check_out_jkt = check_out_utc.astimezone(jakarta_tz)
        else:
            check_out_jkt = None
        
        # Determine status
        status = "Selesai" if log.check_out_time else "Sedang Berkunjung"
        
        history.append({
            "id": log.id,
            "date": check_in_jkt.strftime("%d/%m/%Y") if check_in_jkt else "-",
            "check_in": check_in_jkt.strftime("%H:%M:%S") if check_in_jkt else "-",
            "check_out": check_out_jkt.strftime("%H:%M:%S") if check_out_jkt else None,
            "status": status
        })
    
    return {"history": history}

# === SECURE FILE ACCESS ===
@app.get("/uploads/{filename}", tags=["Files"])
def get_uploaded_file(filename: str, current_admin: models.Admin = Depends(get_current_admin)):
    """
    Protected endpoint for accessing uploaded files.
    Requires admin authentication to prevent unauthorized access to KTP/visitor photos.
    """
    file_path = os.path.join(UPLOAD_DIR, filename)
    
    # Security: Prevent directory traversal
    if ".." in filename or "/" in filename or "\\" in filename:
        logger.warning(f"Directory traversal attempt blocked: {filename}")
        raise HTTPException(status_code=400, detail="Invalid filename")
    
    # Check if file exists
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")
    
    # Security: Force download to prevent XSS (especially for PDFs/SVGs)
    # Use inline for images to display in browser
    media_type = None
    disposition = "inline"  # For images, show in browser
    
    # Detect file type
    file_ext = filename.lower().split('.')[-1]
    if file_ext in ['jpg', 'jpeg', 'png', 'gif']:
        media_type = f"image/{file_ext}" if file_ext != 'jpg' else "image/jpeg"
        disposition = "inline"  # Display images
    elif file_ext == 'pdf':
        media_type = "application/pdf"
        disposition = "attachment"  # Force download PDFs
    
    headers = {
        "Content-Disposition": f"{disposition}; filename={filename}",
        "X-Content-Type-Options": "nosniff",  # Prevent MIME sniffing
        "Cache-Control": "private, max-age=3600"  # Cache for 1 hour
    }
    
    logger.info(f"File access: {filename} by admin {current_admin.username}")
    return FileResponse(file_path, media_type=media_type, headers=headers)

@app.get("/uploads/task_letters/{filename}", tags=["Files"])
def get_task_letter(filename: str, current_admin: models.Admin = Depends(get_current_admin)):
    """
    Protected endpoint for accessing task letters.
    Requires admin authentication.
    """
    file_path = os.path.join(TASK_LETTER_DIR, filename)
    
    # Security: Prevent directory traversal
    if ".." in filename or "/" in filename or "\\" in filename:
        logger.warning(f"Directory traversal attempt blocked: {filename}")
        raise HTTPException(status_code=400, detail="Invalid filename")
    
    # Check if file exists
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")
    
    # Force download/preview
    # PDFs should usually be previewable in modern browsers (inline)
    # but strictly speaking user might want to download. 
    # Let's default to inline for convenience, browsers handle it well.
    
    media_type = "application/pdf"
    if filename.lower().endswith(('.jpg', '.jpeg', '.png')):
        media_type = "image/jpeg"
    
    headers = {
        "Content-Disposition": f"inline; filename={filename}",
        "X-Content-Type-Options": "nosniff",
        "Cache-Control": "private, max-age=3600"
    }
    
    logger.info(f"Task letter access: {filename} by admin {current_admin.username}")
    return FileResponse(file_path, media_type=media_type, headers=headers)

# UPDATE Visitor
@app.put("/visitors/{nik}", tags=["Admin"])
def update_visitor(
    nik: str, 
    full_name: str = Form(...), 
    institution: str = Form(...), 
    phone: str = Form(None),
    current_admin: models.Admin = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    visitor = db.query(models.Visitor).filter(models.Visitor.nik == nik).first()
    if not visitor:
        raise HTTPException(status_code=404, detail="Visitor not found")
    
    visitor.full_name = full_name
    visitor.institution = institution
    visitor.phone = phone
    
    db.commit()
    return {"status": "success", "message": "Data berhasil diperbarui"}

# DELETE Visitor
@app.delete("/visitors/{nik}", tags=["Admin"])
def delete_visitor(
    nik: str, 
    current_admin: models.Admin = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    visitor = db.query(models.Visitor).filter(models.Visitor.nik == nik).first()
    if not visitor:
        raise HTTPException(status_code=404, detail="Visitor not found")
    
    # Hapus log kunjungan secara iteratif agar cascade ke TaskLetter berjalan (ORM Level)
    # Ini penting karena bulk delete (.delete()) by-pass cascade ORM!
    logs = db.query(models.VisitLog).filter(models.VisitLog.visitor_nik == nik).all()
    for log in logs:
        db.delete(log)
    
    # Hapus file foto fisik jika ada
    if visitor.photo_path and os.path.exists(visitor.photo_path):
        try:
            os.remove(visitor.photo_path)
        except OSError:
            pass
            
    if visitor.ktp_path and os.path.exists(visitor.ktp_path):
        try:
            os.remove(visitor.ktp_path)
        except OSError:
            pass

    if visitor.task_letter_path and os.path.exists(visitor.task_letter_path):
        try:
            os.remove(visitor.task_letter_path)
        except OSError:
            pass
        
    db.delete(visitor)
    db.commit()
    return {"status": "success", "message": "Data pengunjung berhasil dihapus"}

# === DOCUMENT MANAGEMENT ENDPOINTS ===

# 1. UPDATE PHOTO
@app.put("/visitors/{nik}/photo", tags=["Admin"])
def update_visitor_photo(
    nik: str,
    photo: UploadFile = File(...),
    current_admin: models.Admin = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    visitor = db.query(models.Visitor).filter(models.Visitor.nik == nik).first()
    if not visitor:
        raise HTTPException(status_code=404, detail="Visitor not found")
    
    # Delete old photo
    if visitor.photo_path and os.path.exists(visitor.photo_path):
        try:
            os.remove(visitor.photo_path)
            logger.info(f"Deleted old photo for {nik}")
        except OSError as e:
            logger.warning(f"Failed to delete old photo: {e}")

    # Save new photo
    new_path = validate_and_save_file(photo)
    visitor.photo_path = new_path
    db.commit()
    
    return {"status": "success", "message": "Foto profil berhasil diperbarui", "path": new_path}

# 2. UPDATE KTP
@app.put("/visitors/{nik}/ktp", tags=["Admin"])
def update_visitor_ktp(
    nik: str,
    ktp: UploadFile = File(...),
    current_admin: models.Admin = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    visitor = db.query(models.Visitor).filter(models.Visitor.nik == nik).first()
    if not visitor:
        raise HTTPException(status_code=404, detail="Visitor not found")
    
    # Delete old KTP
    if visitor.ktp_path and os.path.exists(visitor.ktp_path):
        try:
            os.remove(visitor.ktp_path)
        except OSError:
            pass

    # Save new KTP
    new_path = validate_and_save_file(ktp)
    visitor.ktp_path = new_path
    db.commit()
    
    return {"status": "success", "message": "Foto KTP berhasil diperbarui", "path": new_path}

# 3. GET TASK LETTERS
@app.get("/visitors/{nik}/task-letters", tags=["Admin"])
def get_task_letters(nik: str, db: Session = Depends(get_db)):
    visitor = db.query(models.Visitor).filter(models.Visitor.nik == nik).first()
    if not visitor:
        raise HTTPException(status_code=404, detail="Visitor not found")
        
    documents = []
    
    # A. Legacy Task Letter (from Visitor table)
    if visitor.task_letter_path:
        documents.append({
            "id": "legacy",
            "type": "legacy",
            "filename": os.path.basename(visitor.task_letter_path),
            "stored_filename": os.path.basename(visitor.task_letter_path),
            "uploaded_at": visitor.created_at,
            "path": visitor.task_letter_path
        })
        
    # B. Additional Task Letters (from TaskLetter table)
    # Join with VisitLog to allow filtering/showing visit info if needed
    extra_letters = db.query(models.TaskLetter).join(models.VisitLog)\
        .filter(models.VisitLog.visitor_nik == nik)\
        .order_by(models.TaskLetter.uploaded_at.desc()).all()
        
    for letter in extra_letters:
        documents.append({
            "id": letter.id,
            "type": "additional",
            "filename": letter.original_filename,
            "stored_filename": os.path.basename(letter.file_path),
            "uploaded_at": letter.uploaded_at,
            "path": letter.file_path,
            "visit_date": letter.visit.visit_date
        })
        
    return {"documents": documents}

# 4. UPLOAD TASK LETTER (Manual)
@app.post("/visitors/{nik}/task-letters", tags=["Admin"])
def upload_task_letter(
    nik: str,
    file: UploadFile = File(...),
    current_admin: models.Admin = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    # Find latest active visit or just latest visit
    # Ideally attach to the most recent visit
    latest_visit = db.query(models.VisitLog)\
        .filter(models.VisitLog.visitor_nik == nik)\
        .order_by(models.VisitLog.check_in_time.desc())\
        .first()
        
    if not latest_visit:
        raise HTTPException(400, "Visitor belum memiliki riwayat kunjungan. Tidak bisa upload surat tugas.")

    file_path = validate_and_save_file(file)
    
    new_letter = models.TaskLetter(
        visit_id=latest_visit.id,
        file_path=file_path,
        original_filename=file.filename,
        file_size=0 # Optional implementation to get size
    )
    db.add(new_letter)
    db.commit()
    
    return {"status": "success", "message": "Surat tugas berhasil ditambahkan"}

# 5. DELETE TASK LETTER
@app.delete("/visitors/{nik}/task-letters/{type}/{id}", tags=["Admin"])
def delete_task_letter(
    nik: str, 
    type: str, 
    id: str,
    current_admin: models.Admin = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    visitor = db.query(models.Visitor).filter(models.Visitor.nik == nik).first()
    if not visitor:
        raise HTTPException(404, "Visitor not found")

    if type == "legacy":
        if visitor.task_letter_path:
            if os.path.exists(visitor.task_letter_path):
                try: os.remove(visitor.task_letter_path)
                except: pass
            visitor.task_letter_path = None
            db.commit()
            return {"status": "success", "message": "Legacy task letter deleted"}
        else:
            raise HTTPException(404, "No legacy task letter found")
            
    elif type == "additional":
        letter = db.query(models.TaskLetter).filter(models.TaskLetter.id == int(id)).first()
        if not letter:
            raise HTTPException(404, "Task letter not found")
            
        if os.path.exists(letter.file_path):
            try: os.remove(letter.file_path)
            except: pass
            
        db.delete(letter)
        db.commit()
        return {"status": "success", "message": "Task letter deleted"}
        
    else:
        raise HTTPException(400, "Invalid type")

# FORCE CHECK-OUT (Admin Manual)
@app.put("/admin/visits/{visit_id}/checkout", tags=["Admin"])
def force_checkout_visit(
    visit_id: int,
    current_admin: models.Admin = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Force check-out a visitor manually (Admin only).
    Useful when a visitor forgets to check out.
    """
    visit = db.query(models.VisitLog).filter(models.VisitLog.id == visit_id).first()
    if not visit:
        raise HTTPException(status_code=404, detail="Data kunjungan tidak ditemukan")
    
    if visit.check_out_time:
        raise HTTPException(status_code=400, detail="Pengunjung ini sudah check-out")

    # Set check-out time to NOW (UTC)
    now_jkt = datetime.now(JAKARTA_TZ)
    now_utc = now_jkt.astimezone(pytz.UTC).replace(tzinfo=None)
    
    visit.check_out_time = now_utc
    db.commit()
    
    logger.info(f"Force Check-out by Admin {current_admin.username}: Visit ID {visit_id} at {now_jkt}")
    return {"status": "success", "message": "Berhasil check-out manual", "time": now_jkt.strftime("%H:%M:%S")}

# --- 7. ANALYTICS ENDPOINT ---
@app.get("/analytics/dashboard", dependencies=[Depends(get_current_admin)])
def get_analytics(days: int = 30, db: Session = Depends(get_db)):
    """
    Get aggregated data for analytics dashboard.
    Supports filtering by last N days (default 30).
    """
    jakarta_tz = pytz.timezone('Asia/Jakarta')
    
    # Calculate cutoff date
    cutoff_date = datetime.now() - timedelta(days=days)
    
    # Fetch logs within range (or all for heatmap? usually analytics honors the filter)
    # Let's apply filter to ALL charts for consistency
    logs = db.query(models.VisitLog).filter(models.VisitLog.check_in_time >= cutoff_date).all()
    
    # --- 1. Heatmap (Python Aggregation for Timezone Accuracy) ---
    # Initialize hours 0-23
    hourly_counts = {i: 0 for i in range(24)}
    
    for log in logs:
        if log.check_in_time:
            # Convert UTC (DB) -> Jakarta
            utc_time = log.check_in_time.replace(tzinfo=pytz.UTC)
            jkt_time = utc_time.astimezone(jakarta_tz)
            hourly_counts[jkt_time.hour] += 1
            
    heatmap_data = [{"hour": h, "count": c} for h, c in hourly_counts.items()]
    
    # --- 2. Trend (Daily Counts) ---
    daily_counts = {}
    
    # Iterate logs to count per day
    for log in logs:
        if log.check_in_time:
            utc_time = log.check_in_time.replace(tzinfo=pytz.UTC)
            jkt_time = utc_time.astimezone(jakarta_tz)
            date_str = jkt_time.strftime("%Y-%m-%d")
            daily_counts[date_str] = daily_counts.get(date_str, 0) + 1
            
    # Fill missing dates in range
    trend_data = []
    for i in range(days):
        d = (datetime.now(jakarta_tz) - timedelta(days=days - 1 - i)).strftime("%Y-%m-%d")
        trend_data.append({"date": d, "count": daily_counts.get(d, 0)})

    # --- 3. Top Institutions (Filtered) ---
    institution_counts = {}
    for log in logs:
        # We need visitor data. This is inefficient N+1 if lazy loading.
        # But 'logs' has visitor_nik.
        # Let's optimize: query Visitor info separately or join?
        # Since we already fetched logs, let's use a JOIN query for this part instead of Python loop if logs is large.
        # But for consistency with 'logs' list, let's just grab visitor info.
        # Actually, let's re-query for Top Inst to be safe and efficient.
        pass
        
    top_institutions = db.query(
        models.Visitor.institution,
        func.count(models.Visitor.nik).label('count')
    ).join(models.VisitLog, models.Visitor.nik == models.VisitLog.visitor_nik)\
     .filter(models.VisitLog.check_in_time >= cutoff_date)\
     .group_by(models.Visitor.institution)\
     .order_by(desc('count'))\
     .limit(5).all()
    
    institution_data = [{"name": inst, "count": c} for inst, c in top_institutions]

    # --- 4. Summary Cards ---
    total_visits = db.query(func.count(models.VisitLog.id)).scalar() # All time total
    filtered_visits = len(logs) # Visits in range
    
    today = datetime.now(jakarta_tz).date()
    # Count today's visits (needs DB query or filter 'logs' if 'days' covers today)
    # Safer to query DB for exact "Today" regardless of filter
    # But wait, date comparision in DB (UTC) vs Local.
    # Let's use Python on recent logs if efficient, or DB date func.
    # DB func is risky for timezone. Let's filter 'logs'.
    # If filter is 7 days, today is included.
    
    visits_today = 0
    now_utc = datetime.now(pytz.UTC)
    start_of_day_utc = datetime.now(jakarta_tz).replace(hour=0, minute=0, second=0, microsecond=0).astimezone(pytz.UTC)
    
    visits_today = db.query(func.count(models.VisitLog.id)).filter(models.VisitLog.check_in_time >= start_of_day_utc).scalar()

    return {
        "heatmap": heatmap_data,
        "trend": trend_data,
        "institutions": institution_data,
        "summary": {
            "total_visits": total_visits, # Keep global total
            "visits_in_range": filtered_visits,
            "visits_today": visits_today
        }
    }


# ============================================
# === MASTER DATA: RUANGAN (ROOMS) ===
# ============================================

@app.get("/rooms", tags=["Master Data"])
def get_active_rooms(db: Session = Depends(get_db)):
    """Get all active rooms for dropdown (Public endpoint)"""
    rooms = db.query(models.Room).filter(models.Room.is_active == True).all()
    return [{"id": r.id, "name": r.name, "description": r.description} for r in rooms]

@app.get("/admin/rooms", tags=["Admin - Master Data"])
def get_all_rooms(
    current_admin: models.Admin = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Get all rooms including inactive ones (Admin only)"""
    rooms = db.query(models.Room).all()
    return [{
        "id": r.id, 
        "name": r.name, 
        "description": r.description,
        "is_active": r.is_active,
        "created_at": r.created_at
    } for r in rooms]

@app.post("/admin/rooms", status_code=status.HTTP_201_CREATED, tags=["Admin - Master Data"])
def create_room(
    name: str = Form(...),
    description: str = Form(None),
    current_admin: models.Admin = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Create a new room"""
    # Check if room name already exists
    existing = db.query(models.Room).filter(models.Room.name == name.strip()).first()
    if existing:
        raise HTTPException(400, f"Ruangan '{name}' sudah ada.")
    
    new_room = models.Room(
        name=name.strip(),
        description=description.strip() if description else None
    )
    db.add(new_room)
    db.commit()
    db.refresh(new_room)
    
    logger.info(f"Room created: {name} by {current_admin.username}")
    return {"status": "success", "message": f"Ruangan '{name}' berhasil ditambahkan", "id": new_room.id}

@app.put("/admin/rooms/{room_id}", tags=["Admin - Master Data"])
def update_room(
    room_id: int,
    name: str = Form(None),
    description: str = Form(None),
    is_active: bool = Form(None),
    current_admin: models.Admin = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Update a room"""
    room = db.query(models.Room).filter(models.Room.id == room_id).first()
    if not room:
        raise HTTPException(404, "Ruangan tidak ditemukan")
    
    if name:
        room.name = name.strip()
    if description is not None:
        room.description = description.strip() if description else None
    if is_active is not None:
        room.is_active = is_active
    
    db.commit()
    logger.info(f"Room updated: ID {room_id} by {current_admin.username}")
    return {"status": "success", "message": "Ruangan berhasil diperbarui"}

@app.delete("/admin/rooms/{room_id}", tags=["Admin - Master Data"])
def delete_room(
    room_id: int,
    current_admin: models.Admin = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Delete a room (soft delete - set inactive)"""
    room = db.query(models.Room).filter(models.Room.id == room_id).first()
    if not room:
        raise HTTPException(404, "Ruangan tidak ditemukan")
    
    # Soft delete - just deactivate
    room.is_active = False
    db.commit()
    
    logger.info(f"Room deactivated: ID {room_id} by {current_admin.username}")
    return {"status": "success", "message": "Ruangan berhasil dinonaktifkan"}


# ============================================
# === MASTER DATA: PENDAMPING (COMPANIONS) ===
# ============================================

@app.get("/companions", tags=["Master Data"])
def get_active_companions(db: Session = Depends(get_db)):
    """Get all active companions for dropdown (Public endpoint)"""
    companions = db.query(models.Companion).filter(models.Companion.is_active == True).all()
    return [{"id": c.id, "name": c.name, "position": c.position} for c in companions]

@app.get("/admin/companions", tags=["Admin - Master Data"])
def get_all_companions(
    current_admin: models.Admin = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Get all companions including inactive ones (Admin only)"""
    companions = db.query(models.Companion).all()
    return [{
        "id": c.id, 
        "name": c.name, 
        "position": c.position,
        "is_active": c.is_active,
        "created_at": c.created_at
    } for c in companions]

@app.post("/admin/companions", status_code=status.HTTP_201_CREATED, tags=["Admin - Master Data"])
def create_companion(
    name: str = Form(...),
    position: str = Form(None),
    current_admin: models.Admin = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Create a new companion"""
    new_companion = models.Companion(
        name=name.strip(),
        position=position.strip() if position else None
    )
    db.add(new_companion)
    db.commit()
    db.refresh(new_companion)
    
    logger.info(f"Companion created: {name} by {current_admin.username}")
    return {"status": "success", "message": f"Pendamping '{name}' berhasil ditambahkan", "id": new_companion.id}

@app.put("/admin/companions/{companion_id}", tags=["Admin - Master Data"])
def update_companion(
    companion_id: int,
    name: str = Form(None),
    position: str = Form(None),
    is_active: bool = Form(None),
    current_admin: models.Admin = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Update a companion"""
    companion = db.query(models.Companion).filter(models.Companion.id == companion_id).first()
    if not companion:
        raise HTTPException(404, "Pendamping tidak ditemukan")
    
    if name:
        companion.name = name.strip()
    if position is not None:
        companion.position = position.strip() if position else None
    if is_active is not None:
        companion.is_active = is_active
    
    db.commit()
    logger.info(f"Companion updated: ID {companion_id} by {current_admin.username}")
    return {"status": "success", "message": "Pendamping berhasil diperbarui"}

@app.delete("/admin/companions/{companion_id}", tags=["Admin - Master Data"])
def delete_companion(
    companion_id: int,
    current_admin: models.Admin = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Delete a companion (soft delete - set inactive)"""
    companion = db.query(models.Companion).filter(models.Companion.id == companion_id).first()
    if not companion:
        raise HTTPException(404, "Pendamping tidak ditemukan")
    
    companion.is_active = False
    db.commit()
    
    logger.info(f"Companion deactivated: ID {companion_id} by {current_admin.username}")
    return {"status": "success", "message": "Pendamping berhasil dinonaktifkan"}


# ============================================
# === VISIT DETAILS & TASK LETTERS ===
# ============================================

@app.get("/visits/{visit_id}", tags=["Visitor"])
def get_visit_details(visit_id: int, db: Session = Depends(get_db)):
    """Get visit details including room, companion, and task letters"""
    visit = db.query(models.VisitLog).filter(models.VisitLog.id == visit_id).first()
    if not visit:
        raise HTTPException(404, "Data kunjungan tidak ditemukan")
    
    jakarta_tz = pytz.timezone('Asia/Jakarta')
    
    # Convert times
    check_in_jkt = None
    check_out_jkt = None
    if visit.check_in_time:
        check_in_utc = visit.check_in_time.replace(tzinfo=pytz.UTC)
        check_in_jkt = check_in_utc.astimezone(jakarta_tz)
    if visit.check_out_time:
        check_out_utc = visit.check_out_time.replace(tzinfo=pytz.UTC)
        check_out_jkt = check_out_utc.astimezone(jakarta_tz)
    
    # Get task letters
    task_letters = []
    for letter in visit.task_letters:
        task_letters.append({
            "id": letter.id,
            "original_filename": letter.original_filename,
            "file_size": letter.file_size,
            "uploaded_at": letter.uploaded_at
        })
    
    return {
        "id": visit.id,
        "visitor_nik": visit.visitor_nik,
        "visit_date": visit.visit_date.isoformat() if visit.visit_date else None,
        "check_in_time": check_in_jkt.isoformat() if check_in_jkt else None,
        "check_out_time": check_out_jkt.isoformat() if check_out_jkt else None,
        "visit_purpose": visit.visit_purpose,
        "room": {"id": visit.room.id, "name": visit.room.name} if visit.room else None,
        "companion": {"id": visit.companion.id, "name": visit.companion.name, "position": visit.companion.position} if visit.companion else None,
        "task_letters": task_letters,
        "status": "Selesai" if visit.check_out_time else "Sedang Berkunjung"
    }

@app.put("/visits/{visit_id}/details", tags=["Visitor"])
def update_visit_details(
    visit_id: int,
    visit_purpose: str = Form(None),
    room_id: int = Form(None),
    companion_id: int = Form(None),
    db: Session = Depends(get_db)
):
    """Update visit details (purpose, room, companion) - can be done during or after check-in"""
    visit = db.query(models.VisitLog).filter(models.VisitLog.id == visit_id).first()
    if not visit:
        raise HTTPException(404, "Data kunjungan tidak ditemukan")
    
    # Validate room_id if provided
    if room_id:
        room = db.query(models.Room).filter(models.Room.id == room_id, models.Room.is_active == True).first()
        if not room:
            raise HTTPException(400, "Ruangan tidak valid atau tidak aktif.")
        visit.room_id = room_id
    
    # Validate companion_id if provided
    if companion_id:
        companion = db.query(models.Companion).filter(models.Companion.id == companion_id, models.Companion.is_active == True).first()
        if not companion:
            raise HTTPException(400, "Pendamping tidak valid atau tidak aktif.")
        visit.companion_id = companion_id
    
    if visit_purpose is not None:
        visit.visit_purpose = visit_purpose.strip() if visit_purpose else None
    
    db.commit()
    logger.info(f"Visit details updated: ID {visit_id}")
    return {"status": "success", "message": "Detail kunjungan berhasil diperbarui"}


# === TASK LETTERS (SURAT TUGAS) ===

def validate_pdf_file(upload_file: UploadFile) -> str:
    """Validate and save PDF file for task letter"""
    if not upload_file:
        raise HTTPException(400, "File wajib diupload.")
    
    # Check file extension
    file_ext = os.path.splitext(upload_file.filename)[1].lower()
    if file_ext != ".pdf":
        raise HTTPException(400, "Hanya file PDF yang diizinkan untuk surat tugas.")
    
    # Check file size
    upload_file.file.seek(0, 2)  # Seek to end
    file_size = upload_file.file.tell()
    upload_file.file.seek(0)  # Reset to beginning
    
    if file_size > MAX_FILE_SIZE:
        raise HTTPException(400, f"Ukuran file terlalu besar. Maksimal {MAX_FILE_SIZE // (1024*1024)}MB.")
    
    # Validate MIME type
    try:
        header = upload_file.file.read(2048)
        upload_file.file.seek(0)
        
        detected_mimes = puremagic.magic_string(header)
        if not detected_mimes:
            raise ValueError("Unknown file type")
        mime_type = detected_mimes[0].mime_type
        
        if mime_type != "application/pdf":
            raise HTTPException(400, f"Tipe file tidak valid. Harus PDF (terdeteksi: {mime_type})")
    except Exception as e:
        logger.error(f"Error validating PDF: {e}")
        raise HTTPException(400, "Gagal memvalidasi file PDF.")
    
    # Save file
    unique_name = f"{uuid.uuid4()}{file_ext}"
    file_path = os.path.join(TASK_LETTER_DIR, unique_name)
    
    try:
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(upload_file.file, buffer)
        logger.info(f"Task letter uploaded: {unique_name}")
    except Exception as e:
        logger.error(f"Failed to save task letter: {e}")
        raise HTTPException(500, "Gagal menyimpan file surat tugas.")
    
    return file_path, file_size

@app.post("/visits/{visit_id}/task-letters", status_code=status.HTTP_201_CREATED, tags=["Visitor"])
def upload_task_letter(
    visit_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """Upload a task letter (surat tugas) PDF for a visit"""
    visit = db.query(models.VisitLog).filter(models.VisitLog.id == visit_id).first()
    if not visit:
        raise HTTPException(404, "Data kunjungan tidak ditemukan")
    
    # Check if already at max limit
    current_count = db.query(models.TaskLetter).filter(models.TaskLetter.visit_id == visit_id).count()
    if current_count >= MAX_TASK_LETTERS_PER_VISIT:
        raise HTTPException(400, f"Maksimal {MAX_TASK_LETTERS_PER_VISIT} file surat tugas per kunjungan.")
    
    # Validate and save file
    file_path, file_size = validate_pdf_file(file)
    
    # Create task letter record
    new_letter = models.TaskLetter(
        visit_id=visit_id,
        file_path=file_path,
        original_filename=file.filename,
        file_size=file_size
    )
    db.add(new_letter)
    db.commit()
    db.refresh(new_letter)
    
    logger.info(f"Task letter uploaded for visit {visit_id}: {file.filename}")
    return {
        "status": "success", 
        "message": "Surat tugas berhasil diupload",
        "id": new_letter.id,
    }

@app.get("/visitors/{nik}/task-letters", tags=["Visitor"])
def get_task_letters_for_visitor(nik: str, db: Session = Depends(get_db)):
    """Get task letters for visitor's CURRENT active visit only (prevents mixing between visits)"""
    # Find visitor
    visitor = db.query(models.Visitor).filter(models.Visitor.nik == nik).first()
    if not visitor:
        raise HTTPException(404, "Visitor tidak ditemukan")
    
    # Find ACTIVE visit (checked in today, not checked out yet)
    jakarta_tz = pytz.timezone('Asia/Jakarta')
    now_jkt = datetime.now(jakarta_tz)
    
    active_visit = db.query(models.VisitLog).filter(
        models.VisitLog.visitor_nik == nik,
        models.VisitLog.visit_date == now_jkt.date(),
        models.VisitLog.check_out_time == None
    ).first()
    
    if not active_visit:
        # No active visit - return empty documents
        return {"documents": [], "visit_id": None, "status": "no_active_visit"}
    
    # Get task letters ONLY for this specific visit
    letters = db.query(models.TaskLetter).filter(
        models.TaskLetter.visit_id == active_visit.id
    ).all()
    
    documents = [{
        "id": l.id,
        "filename": l.original_filename,
        "stored_filename": os.path.basename(l.file_path),
        "file_size": l.file_size,
        "uploaded_at": l.uploaded_at.isoformat() if l.uploaded_at else None,
        "type": "additional",
        "visit_date": str(active_visit.visit_date)
    } for l in letters]
    
    return {
        "documents": documents, 
        "visit_id": active_visit.id,
        "status": "active_visit"
    }


@app.get("/visits/{visit_id}/task-letters", tags=["Visitor"])
def get_task_letters(visit_id: int, db: Session = Depends(get_db)):
    """Get all task letters for a visit"""
    visit = db.query(models.VisitLog).filter(models.VisitLog.id == visit_id).first()
    if not visit:
        raise HTTPException(404, "Data kunjungan tidak ditemukan")
    
    letters = db.query(models.TaskLetter).filter(models.TaskLetter.visit_id == visit_id).all()
    
    return {
        "visit_id": visit_id,
        "count": len(letters),
        "max_allowed": MAX_TASK_LETTERS_PER_VISIT,
        "task_letters": [{
            "id": l.id,
            "original_filename": l.original_filename,
            "file_size": l.file_size,
            "uploaded_at": l.uploaded_at.isoformat() if l.uploaded_at else None
        } for l in letters]
    }

@app.get("/visits/{visit_id}/task-letters/{letter_id}/download", tags=["Visitor"])
def download_task_letter(visit_id: int, letter_id: int, db: Session = Depends(get_db)):
    """Download a specific task letter"""
    letter = db.query(models.TaskLetter).filter(
        models.TaskLetter.id == letter_id,
        models.TaskLetter.visit_id == visit_id
    ).first()
    
    if not letter:
        raise HTTPException(404, "Surat tugas tidak ditemukan")
    
    if not os.path.exists(letter.file_path):
        logger.error(f"Task letter file missing: {letter.file_path}")
        raise HTTPException(404, "File surat tugas tidak ditemukan di server")
    
    return FileResponse(
        letter.file_path,
        filename=letter.original_filename,
        media_type="application/pdf"
    )

@app.get("/visits/{visit_id}/task-letters/archive", tags=["Visitor"])
def download_task_letters_archive(visit_id: int, db: Session = Depends(get_db)):
    """Convert multiple task letters into a single ZIP file for download"""
    visit = db.query(models.VisitLog).filter(models.VisitLog.id == visit_id).first()
    if not visit:
        raise HTTPException(404, "Data kunjungan tidak ditemukan")
    
    letters = db.query(models.TaskLetter).filter(models.TaskLetter.visit_id == visit_id).all()
    if not letters:
        raise HTTPException(404, "Tidak ada surat tugas untuk kunjungan ini")
        
    # Create in-memory ZIP
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
        for letter in letters:
            if os.path.exists(letter.file_path):
                # Ensure unique arcnames in zip
                arcname = letter.original_filename
                # Simple duplicate check - in production might need more robust handling
                zip_file.write(letter.file_path, arcname=arcname)
            else:
                logger.warning(f"Skipping missing file in ZIP: {letter.file_path}")
    
    zip_buffer.seek(0)
    
    filename = f"Surat_Tugas_{visit.visitor_nik}_{visit.visit_date}.zip"
    
    return Response(
        content=zip_buffer.getvalue(),
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.delete("/visits/{visit_id}/task-letters/{letter_id}", tags=["Visitor"])
def delete_task_letter(visit_id: int, letter_id: int, db: Session = Depends(get_db)):
    """Delete a task letter"""
    letter = db.query(models.TaskLetter).filter(
        models.TaskLetter.id == letter_id,
        models.TaskLetter.visit_id == visit_id
    ).first()
    
    if not letter:
        raise HTTPException(404, "Surat tugas tidak ditemukan")
    
    # Delete physical file
    if os.path.exists(letter.file_path):
        try:
            os.remove(letter.file_path)
        except Exception as e:
            logger.error(f"Failed to delete task letter file: {e}")
    
    # Delete database record
    db.delete(letter)
    db.commit()
    
    logger.info(f"Task letter deleted: ID {letter_id} from visit {visit_id}")
    return {"status": "success", "message": "Surat tugas berhasil dihapus"}


# === GET CURRENT ACTIVE VISIT ===
@app.get("/visitors/{nik}/active-visit", tags=["Visitor"])
def get_active_visit(nik: str, db: Session = Depends(get_db)):
    """Get current active visit for a visitor (if any)"""
    now_jkt = datetime.now(JAKARTA_TZ)
    
    active_visit = db.query(models.VisitLog).filter(
        models.VisitLog.visitor_nik == nik,
        models.VisitLog.visit_date == now_jkt.date(),
        models.VisitLog.check_out_time == None
    ).first()
    
    if not active_visit:
        return {"has_active_visit": False, "visit": None}
    
    # Get details
    task_letters_count = db.query(models.TaskLetter).filter(
        models.TaskLetter.visit_id == active_visit.id
    ).count()
    
    return {
        "has_active_visit": True,
        "visit": {
            "id": active_visit.id,
            "visit_purpose": active_visit.visit_purpose,
            "room": {"id": active_visit.room.id, "name": active_visit.room.name} if active_visit.room else None,
            "companion": {"id": active_visit.companion.id, "name": active_visit.companion.name} if active_visit.companion else None,
            "task_letters_count": task_letters_count,
            "can_add_more_letters": task_letters_count < MAX_TASK_LETTERS_PER_VISIT
        }
    }

# === FILE SERVING ENDPOINT ===
@app.get("/uploads/{filename}", tags=["System"])
def get_uploaded_file(filename: str):
    """
    Serve uploaded files (photos, PDFs, etc).
    """
    # 1. Search in main UPLOAD_DIR
    file_path = os.path.join(UPLOAD_DIR, filename)
    if os.path.exists(file_path):
        return FileResponse(file_path)
    
    # 2. Search in TASK_LETTER_DIR
    task_letter_path = os.path.join(TASK_LETTER_DIR, filename)
    if os.path.exists(task_letter_path):
        return FileResponse(task_letter_path)
    
    raise HTTPException(status_code=404, detail="File not found")

